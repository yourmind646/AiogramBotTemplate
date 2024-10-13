# Aiogram
import aiogram.types as types
from aiogram.fsm.context import FSMContext
from aiogram.filters import StateFilter
from aiogram import Router, F

# DB
from database.orm import ORM as orm

# States
from states.main_states import AdminStates

# Another
import shutil, os
from openpyxl import load_workbook

# Init
list_of_users_router = Router()


@list_of_users_router.message(F.text == "📑 Список пользователей", StateFilter(AdminStates.main))
async def process_list_of_users(message: types.Message, state: FSMContext):

	# copy the table
	table_path = shutil.copy(src = "templates/users.xlsx", dst = f"templates/users_list.xlsx")

	# load table
	book = load_workbook(filename = table_path)
	sheet = book["users"]

	# get all users
	all_clients = await orm.get_all_users()

	# write
	for row, user in enumerate(all_clients, 2):
		# write row
		sheet.cell(row = row, column = 1, value = user.user_id)
		sheet.cell(row = row, column = 2, value = user.username)
		sheet.cell(row = row, column = 3, value = user.fullname)
		sheet.cell(row = row, column = 4, value = user.register_date.strftime(r"%d-%m-%y %H:%M %Z%z"))

	# save book
	book.save(table_path)

	# answer
	await message.answer_document(document = types.FSInputFile(table_path))

	# delete temp table
	try:
		os.remove(table_path)
	except:
		pass
